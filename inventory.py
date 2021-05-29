import openpyxl as xl


def read_products_from(file_name, sheet_name):
    inventory_file = xl.load_workbook(file_name)
    return inventory_file[sheet_name]


def group_products_by_supplier(products):
    products_per_all_suppliers = {}
    for row in range(2, products.max_row + 1):
        supplier = products.cell(row, 4).value
        if supplier in products_per_all_suppliers.keys():
            products_per_all_suppliers[supplier] += 1
        else:
            products_per_all_suppliers[supplier] = 1
    return products_per_all_suppliers
